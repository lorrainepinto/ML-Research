import openpyxl
import os
from openpyxl import Workbook

class Create_sheet:
	def __init__(self, filename ):
		self.filename = filename
	def create(self, path):
		wb = Workbook()
		ws = wb.active
		images = os.listdir(path)
		i = 1
		for img in images:

			ws.cell(i, 1, '0')
			ws.cell(i, 2, path + "//" + img)
			i += 1

		wb.save('Images.xls')

