import openpyxl
import os
from openpyxl import Workbook
from Zernike import ZernikeMoments

path1 = "C:\\Users\\pinto\\Desktop\\DTlabz\\ML\\DHCD\\test\\36"
path2 = "C:\\Users\\pinto\\Desktop\\DTlabz\\ML\\DHCD\\test\\38"

wb = Workbook()
ws = wb.active
ws.title = "RA"
images = os.listdir(path1)

desc = ZernikeMoments(10)
row = 1
for img in images:
    moments = desc.describe(path1+"\\"+img)
    col = 1
    ws.cell(row, col, '0')   # '1' for path2 (it's a label)
    for val in moments:
        col += 1
        ws.cell(row, col, val)
    row += 1

ws1 = wb.create_sheet("WA")
images = os.listdir(path2)
row = 1
for img in images:
    moments = desc.describe(path2+"\\"+img)
    col = 1
    ws1.cell(row, col, '1')   # '1' for path2 (it's a label)
    for val in moments:
        col += 1
        ws1.cell(row, col, val)
    row += 1

wb.save('Both.xls')
